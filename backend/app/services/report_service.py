"""Report generation — Excel + PDF exports."""

from __future__ import annotations

import io
from datetime import datetime, timezone
from typing import Iterable

from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy.orm import Session

from app.models.department import Department
from app.models.student import Student
from app.repositories.prediction_repo import prediction_repo
from app.repositories.student_repo import student_repo


def students_xlsx(db: Session) -> bytes:
    """Return an Excel workbook of all students + their latest risk."""
    wb = Workbook()
    ws = wb.active
    ws.title = "students"
    ws.append(
        [
            "Roll No", "Name", "Department", "Semester", "Attendance %",
            "Internal", "Semester", "Backlogs", "Fee Paid", "Fee Delay (days)",
            "Latest Risk", "Confidence", "Updated",
        ]
    )

    students, _ = student_repo.search(
        db, q=None, department_id=None, risk=None, page=1, page_size=10**6, sort="-created_at"
    )
    for s in students:
        latest = prediction_repo.latest(db, s.id)
        ws.append(
            [
                s.roll_no, s.name,
                s.department.code if s.department else "",
                s.semester, s.attendance_pct, s.internal_marks, s.semester_marks,
                s.backlogs, "Yes" if s.fee_paid else "No", s.fee_delay_days,
                latest.risk_level.value if latest else "",
                round(float(latest.confidence), 3) if latest else "",
                s.updated_at.isoformat() if s.updated_at else "",
            ]
        )

    for col_idx in range(1, ws.max_column + 1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        max_len = max(
            (len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(1, ws.max_row + 1)),
            default=10,
        )
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def student_pdf(db: Session, student_id: int) -> bytes | None:
    s = student_repo.get(db, student_id)
    if not s:
        return None
    latest = prediction_repo.latest(db, s.id)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, title=f"Dropout Risk Report — {s.roll_no}")
    styles = getSampleStyleSheet()
    flow = []
    flow.append(Paragraph(f"<b>Dropout Risk Report — {s.name} ({s.roll_no})</b>", styles["Title"]))
    flow.append(Paragraph(datetime.now(timezone.utc).strftime("Generated %Y-%m-%d %H:%M UTC"), styles["Italic"]))
    flow.append(Spacer(1, 12))

    flow.append(Paragraph("<b>Student Profile</b>", styles["Heading2"]))
    flow.append(_kv_table([
        ("Department", s.department.code if s.department else ""),
        ("Semester", str(s.semester)),
        ("Age / Gender", f"{s.age} / {s.gender}"),
        ("Attendance %", f"{s.attendance_pct:.1f}"),
        ("Internal marks", f"{s.internal_marks:.1f}"),
        ("Semester marks", f"{s.semester_marks:.1f}"),
        ("Backlogs", str(s.backlogs)),
        ("Fee paid", "Yes" if s.fee_paid else "No"),
        ("Fee delay (days)", str(s.fee_delay_days)),
        ("Financial status", s.financial_status.value),
        ("Placement readiness", s.placement_readiness.value),
    ]))
    flow.append(Spacer(1, 12))

    flow.append(Paragraph("<b>Latest Prediction</b>", styles["Heading2"]))
    if latest:
        flow.append(_kv_table([
            ("Risk level", latest.risk_level.value.upper()),
            ("Confidence", f"{latest.confidence:.2f}"),
            ("Model", latest.model_version),
            ("Generated", latest.created_at.strftime("%Y-%m-%d %H:%M")),
        ]))
        narrative = (latest.explanation_json or {}).get("narrative", "")
        if narrative:
            flow.append(Spacer(1, 8))
            flow.append(Paragraph(f"<i>{narrative}</i>", styles["BodyText"]))
        factors = (latest.explanation_json or {}).get("top_factors", [])
        if factors:
            flow.append(Spacer(1, 8))
            data = [["Feature", "Value", "Direction", "Contribution"]]
            for f in factors:
                data.append([f.get("feature"), str(f.get("value")), f.get("direction"), f"{f.get('contribution'):.3f}"])
            t = Table(data, hAlign="LEFT")
            t.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
            ]))
            flow.append(t)
    else:
        flow.append(Paragraph("No prediction has been generated for this student yet.", styles["BodyText"]))

    doc.build(flow)
    return buf.getvalue()


def department_pdf(db: Session, department_id: int) -> bytes | None:
    dept = db.get(Department, department_id)
    if not dept:
        return None
    students, _ = student_repo.search(
        db, q=None, department_id=department_id, risk=None, page=1, page_size=10**6, sort="roll_no"
    )

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, title=f"Department Report — {dept.code}")
    styles = getSampleStyleSheet()
    flow = []
    flow.append(Paragraph(f"<b>Department Report — {dept.name} ({dept.code})</b>", styles["Title"]))
    flow.append(Paragraph(datetime.now(timezone.utc).strftime("Generated %Y-%m-%d %H:%M UTC"), styles["Italic"]))
    flow.append(Spacer(1, 12))
    flow.append(Paragraph(f"Total students: <b>{len(students)}</b>", styles["BodyText"]))
    flow.append(Spacer(1, 12))

    data = [["Roll", "Name", "Sem", "Attend %", "Internal", "Backlogs", "Risk"]]
    for s in students:
        latest = prediction_repo.latest(db, s.id)
        data.append([
            s.roll_no, s.name, str(s.semester), f"{s.attendance_pct:.0f}",
            f"{s.internal_marks:.0f}", str(s.backlogs),
            latest.risk_level.value if latest else "—",
        ])
    table = Table(data, repeatRows=1, hAlign="LEFT")
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#dbe2ff")),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (2, 1), (-1, -1), "CENTER"),
    ]))
    flow.append(table)
    doc.build(flow)
    return buf.getvalue()


def _kv_table(rows: Iterable[tuple[str, str]]) -> Table:
    data = [[Paragraph(f"<b>{k}</b>", getSampleStyleSheet()["BodyText"]), v] for k, v in rows]
    t = Table(data, colWidths=[150, 320], hAlign="LEFT")
    t.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    return t
